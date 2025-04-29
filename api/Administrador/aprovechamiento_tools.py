# 📄 descargar_plantilla_aprovechamiento con formato profesional

import pandas as pd
from django.http import HttpResponse
from api.models import ProgramaEducativoAntiguo, ProgramaEducativoNuevo, CicloPeriodo
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def descargar_plantilla_aprovechamiento(request):
    # Obtener programas educativos
    programas_antiguos = list(ProgramaEducativoAntiguo.objects.all())
    programas_nuevos = list(ProgramaEducativoNuevo.objects.all())

    # Obtener ciclos-periodo
    ciclos = CicloPeriodo.objects.select_related('ciclo', 'periodo').order_by('ciclo__anio', 'periodo__clave')
    columnas = [f"{cp.periodo.clave} {cp.ciclo.anio}" for cp in ciclos]
    columnas.insert(0, 'Programa Educativo')

    # Crear datos
    datos = []
    for prog in programas_antiguos + programas_nuevos:
        fila = [prog.nombre] + ['' for _ in ciclos]
        datos.append(fila)

    df = pd.DataFrame(datos, columns=columnas)

    # Crear workbook openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Aprovechamiento"

    # Escribir DataFrame en el Excel
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Formato de encabezados
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    # Formato general de celdas
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Ancho de columnas
    ws.column_dimensions['A'].width = 40  # Programa Educativo
    for col in range(2, len(columnas) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15

    # Definir respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="PLANTILLA_APROVECHAMIENTO_FORMATO.xlsx"'

    wb.save(response)
    return response
