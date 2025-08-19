# api/Administrador/titulados_tsu_inge_view.py
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.views.decorators.http import require_GET, require_POST
from django.db import transaction
from django.db.models import F, Q, Value as V
from django.db.models.functions import Coalesce
from django.core.paginator import Paginator

from api.models import (
    ProgramaEducativoAntiguo,
    ProgramaEducativoNuevo,
    ProgramaEducativo,   # catálogo (id, tipo) (fallback)
    TituladosHistoricos,
)

# ====== Excel (openpyxl) ======
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ----------------- helpers -----------------
def _y(x):
    try:
        return int(float(x))
    except Exception:
        return 0

def _has_field(model, name: str) -> bool:
    return any(f.name == name for f in model._meta.get_fields())

def _int_pos(v, d=0):
    if v in (None, ""):
        return d
    try:
        x = int(float(v))
        return x if x >= 0 else d
    except Exception:
        return d

def _sem(v):
    # default 10; solo 5 o 10
    x = _int_pos(v, 10) or 10
    return 5 if x == 5 else 10

# Encabezados de la plantilla (NO CAMBIAR)
CAMPOS_PLANTILLA = [
    "anio_ingreso",
    "anio_egreso",
    "titulados_hombres",
    "titulados_mujeres",
    "registrados_dgp_h",
    "registrados_dgp_m",
    "programa_antiguo_id",   # FK por id (char)
    "programa_nuevo_id",     # FK por id (char)
    "semestre",              # 5 TSU / 10 Ing (se sobrescribe en carga por endpoint)
]

# ------------- lógica de filtros compartida (para la vista y export) -------------
def _filtrar_queryset_base(request):
    """Devuelve (nivel, qs, qs_named) aplicando los mismos filtros que la vista."""
    nivel = (request.GET.get("nivel") or "TSU").upper()
    nivel = "TSU" if nivel == "TSU" else "ING"

    qs = (
        TituladosHistoricos.objects
        .select_related("programa_antiguo", "programa_nuevo")
        .order_by("anio_ingreso", "anio_egreso", "programa_antiguo__id", "programa_nuevo__id")
    )

    # Filtro por nivel (preferente por campo semestre)
    if _has_field(TituladosHistoricos, "semestre"):
        semestre_objetivo = 5 if nivel == "TSU" else 10
        qs = qs.filter(semestre=semestre_objetivo)
    else:
        # Fallback: catálogo ProgramaEducativo (si lo usas)
        tec_ids = set(
            ProgramaEducativo.objects
            .filter(tipo__iexact="TECNICO")
            .values_list("id", flat=True)
        )
        ing_ids = set(
            ProgramaEducativo.objects
            .filter(tipo__iexact="INGENIERO")
            .values_list("id", flat=True)
        )
        if tec_ids or ing_ids:
            qs = qs.annotate(
                prog_id=Coalesce(
                    F("programa_antiguo__id"),
                    F("programa_nuevo__id"),
                )
            )
            if nivel == "TSU" and tec_ids:
                qs = qs.filter(prog_id__in=tec_ids)
            elif nivel == "ING" and ing_ids:
                qs = qs.filter(prog_id__in=ing_ids)

    # Filtros UI
    anios = request.GET.getlist("anio")
    if anios:
        anios_int = [a for a in (_y(x) for x in anios) if a]
        if anios_int:
            qs = qs.filter(anio_ingreso__in=anios_int)

    programas = request.GET.getlist("programa")
    if programas:
        qs = qs.filter(
            Q(programa_antiguo__nombre__in=programas) |
            Q(programa_nuevo__nombre__in=programas)
        )

    buscar = (request.GET.get("buscar") or "").strip()
    if buscar:
        qs = qs.filter(
            Q(programa_antiguo__nombre__icontains=buscar) |
            Q(programa_nuevo__nombre__icontains=buscar)
        )

    qs_named = qs.annotate(
        programa_nombre=Coalesce(
            F("programa_antiguo__nombre"),
            F("programa_nuevo__nombre"),
            V("SIN PROGRAMA"),
        )
    )
    return nivel, qs, qs_named

# ===== Helper: arma Excel con encabezados, datos y catálogos =====
def _armar_excel_captura_con_catalogos(qs, semestre_por_defecto: int, nombre_archivo: str) -> HttpResponse:
    """
    Crea un Excel con:
      - Hoja 'Captura': encabezados CAMPOS_PLANTILLA + datos de qs
      - Hoja 'Catalogos': programas antiguos y nuevos
      - Validaciones de lista para columnas de FK
      - Si no hay datos, agrega una fila de ejemplo
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Captura"

    # Encabezados
    for col, header in enumerate(CAMPOS_PLANTILLA, start=1):
        ws.cell(row=1, column=col, value=header)

    # Datos
    valores = qs.values(
        "anio_ingreso", "anio_egreso",
        "titulados_hombres", "titulados_mujeres",
        "registrados_dgp_h", "registrados_dgp_m",
        "programa_antiguo_id", "programa_nuevo_id", "semestre",
    )
    r = 2
    for v in valores:
        fila = [
            int(v["anio_ingreso"] or 0),
            int(v["anio_egreso"] or 0),
            int(v["titulados_hombres"] or 0),
            int(v["titulados_mujeres"] or 0),
            int(v["registrados_dgp_h"] or 0),
            int(v["registrados_dgp_m"] or 0),
            v["programa_antiguo_id"] or "",
            v["programa_nuevo_id"] or "",
            int(v["semestre"] or semestre_por_defecto),
        ]
        for c, val in enumerate(fila, start=1):
            ws.cell(row=r, column=c, value=val)
        r += 1

    # Si no hay datos, deja una fila ejemplo
    if r == 2:
        ws.cell(row=2, column=1, value=2020)
        ws.cell(row=2, column=2, value=2023)
        ws.cell(row=2, column=3, value=0)
        ws.cell(row=2, column=4, value=0)
        ws.cell(row=2, column=5, value=0)
        ws.cell(row=2, column=6, value=0)
        ws.cell(row=2, column=7, value="")  # programa_antiguo_id
        ws.cell(row=2, column=8, value="")  # programa_nuevo_id
        ws.cell(row=2, column=9, value=semestre_por_defecto)

    # Hoja Catálogos
    wsc = wb.create_sheet("Catalogos")
    wsc.cell(row=1, column=1, value="programa_antiguo_id")
    wsc.cell(row=1, column=2, value="nombre_antiguo")
    wsc.cell(row=1, column=4, value="programa_nuevo_id")
    wsc.cell(row=1, column=5, value="nombre_nuevo")

    antiguos = list(ProgramaEducativoAntiguo.objects.order_by("id").values_list("id", "nombre"))
    nuevos   = list(ProgramaEducativoNuevo.objects.order_by("id").values_list("id", "nombre"))

    for i, (pid, nom) in enumerate(antiguos, start=2):
        wsc.cell(row=i, column=1, value=pid)
        wsc.cell(row=i, column=2, value=nom)

    for i, (pid, nom) in enumerate(nuevos, start=2):
        wsc.cell(row=i, column=4, value=pid)
        wsc.cell(row=i, column=5, value=nom)

    # Validaciones de lista
    last_ant = (len(antiguos) + 1) if antiguos else 2
    last_nue = (len(nuevos) + 1) if nuevos else 2
    rango_ant = f"Catalogos!$A$2:$A${last_ant}"
    rango_nue = f"Catalogos!$D$2:$D${last_nue}"

    dv_ant = DataValidation(type="list", formula1=rango_ant, allow_blank=True, showDropDown=True)
    dv_nue = DataValidation(type="list", formula1=rango_nue, allow_blank=True, showDropDown=True)
    ws.add_data_validation(dv_ant)
    ws.add_data_validation(dv_nue)

    col_ant = CAMPOS_PLANTILLA.index("programa_antiguo_id") + 1
    col_nue = CAMPOS_PLANTILLA.index("programa_nuevo_id") + 1

    last_row = max(r, 200)
    dv_ant.add(f"{get_column_letter(col_ant)}2:{get_column_letter(col_ant)}{last_row}")
    dv_nue.add(f"{get_column_letter(col_nue)}2:{get_column_letter(col_nue)}{last_row}")

    # Respuesta
    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    wb.save(resp)
    return resp

# ===== Helpers subir: procesar excel parametrizado por semestre destino =====
def _procesar_excel_en_bd(f, semestre_destino: int):
    """
    Lee 'Captura' y hace update_or_create con clave:
      (anio_ingreso, anio_egreso, semestre_destino, programa_antiguo_id, programa_nuevo_id)
    Ignora el valor 'semestre' que venga en el Excel; se fuerza al endpoint.
    """
    try:
        wb = load_workbook(f, data_only=True)
    except Exception as e:
        raise ValueError("El archivo no es un Excel válido.") from e

    if "Captura" not in wb.sheetnames:
        raise ValueError("La hoja 'Captura' no existe. Descarga el Excel desde la página.")

    ws = wb["Captura"]

    # Map headers
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if isinstance(h, str):
            h = h.strip()
        headers[h] = c

    faltan = [h for h in CAMPOS_PLANTILLA if h not in headers]
    if faltan:
        raise ValueError(f"El Excel no tiene las columnas esperadas: {', '.join(faltan)}")

    # Caches de FKs
    antiguos_valid = set(ProgramaEducativoAntiguo.objects.values_list("id", flat=True))
    nuevos_valid   = set(ProgramaEducativoNuevo.objects.values_list("id", flat=True))

    updated, created, skipped = 0, 0, 0
    warnings_list, errors_list = [], []

    with transaction.atomic():
        for r in range(2, ws.max_row + 1):
            # Fila totalmente vacía => continuar
            if all((ws.cell(r, headers[k]).value in (None, "")) for k in CAMPOS_PLANTILLA):
                continue

            data = {k: ws.cell(r, headers[k]).value for k in CAMPOS_PLANTILLA}

            anio_ingreso = _int_pos(data["anio_ingreso"], 0)
            anio_egreso  = _int_pos(data["anio_egreso"], 0)
            tit_h        = _int_pos(data["titulados_hombres"], 0)
            tit_m        = _int_pos(data["titulados_mujeres"], 0)
            dgp_h        = _int_pos(data["registrados_dgp_h"], 0)
            dgp_m        = _int_pos(data["registrados_dgp_m"], 0)

            prog_ant = (str(data["programa_antiguo_id"]).strip() or None) if data["programa_antiguo_id"] not in (None, "") else None
            prog_nue = (str(data["programa_nuevo_id"]).strip() or None) if data["programa_nuevo_id"] not in (None, "") else None

            if not prog_ant and not prog_nue:
                skipped += 1
                warnings_list.append(f"Fila {r}: sin programa_antiguo_id ni programa_nuevo_id. Saltada.")
                continue

            if prog_ant and prog_nue:
                # Priorizamos el NUEVO y avisamos
                warnings_list.append(f"Fila {r}: venían ambos programas; se usó programa_nuevo_id='{prog_nue}'.")
                prog_ant = None

            if prog_ant and prog_ant not in antiguos_valid:
                errors_list.append(f"Fila {r}: programa_antiguo_id '{prog_ant}' no existe.")
                continue
            if prog_nue and prog_nue not in nuevos_valid:
                errors_list.append(f"Fila {r}: programa_nuevo_id '{prog_nue}' no existe.")
                continue

            if anio_ingreso <= 0 or anio_egreso <= 0:
                errors_list.append(f"Fila {r}: años de ingreso/egreso inválidos.")
                continue

            defaults = dict(
                titulados_hombres=tit_h,
                titulados_mujeres=tit_m,
                registrados_dgp_h=dgp_h,
                registrados_dgp_m=dgp_m,
            )
            obj, was_created = TituladosHistoricos.objects.update_or_create(
                anio_ingreso=anio_ingreso,
                anio_egreso=anio_egreso,
                semestre=semestre_destino,            # 👈 forzado por endpoint
                programa_antiguo_id=prog_ant,
                programa_nuevo_id=prog_nue,
                defaults=defaults
            )
            if was_created:
                created += 1
            else:
                updated += 1

    return created, updated, skipped, warnings_list, errors_list

# =================== ENDPOINTS: DESCARGAR (2) ===================
@require_GET
def descargar_plantilla_titulados_tsu(request):
    # TSU = semestre 5; respeta filtros de la vista si los pasas
    _, qs, _ = _filtrar_queryset_base(request)
    qs = qs.filter(semestre=5) if _has_field(TituladosHistoricos, "semestre") else qs
    return _armar_excel_captura_con_catalogos(qs, semestre_por_defecto=5, nombre_archivo="titulados_tsu.xlsx")

@require_GET
def descargar_plantilla_titulados_ing(request):
    # Ingeniería = semestre 10
    _, qs, _ = _filtrar_queryset_base(request)
    qs = qs.filter(semestre=10) if _has_field(TituladosHistoricos, "semestre") else qs
    return _armar_excel_captura_con_catalogos(qs, semestre_por_defecto=10, nombre_archivo="titulados_ing.xlsx")

# =================== ENDPOINTS: SUBIR (2) ===================
@require_POST
def subir_titulados_tsu_excel(request):
    f = request.FILES.get("archivo_excel_tsu")
    if not f:
        messages.error(request, "Selecciona un archivo .xlsx para TSU.")
        return redirect("titulados_tsu_inge")
    try:
        c, u, s, warns, errs = _procesar_excel_en_bd(f, semestre_destino=5)
    except ValueError as e:
        messages.error(request, str(e))
        return redirect("titulados_tsu_inge")

    if c: messages.success(request, f"[TSU] Se CREARON {c} registro(s).")
    if u: messages.success(request, f"[TSU] Se ACTUALIZARON {u} registro(s).")
    if s: messages.info(request, f"[TSU] Se saltaron {s} fila(s) vacías o sin programa.")
    if warns: messages.warning(request, " · ".join(warns[:6]) + (" ..." if len(warns) > 6 else ""))
    if errs: messages.error(request, " · ".join(errs[:6]) + (" ..." if len(errs) > 6 else ""))
    return redirect("titulados_tsu_inge")

@require_POST
def subir_titulados_ing_excel(request):
    f = request.FILES.get("archivo_excel_ing")
    if not f:
        messages.error(request, "Selecciona un archivo .xlsx para Ingeniería.")
        return redirect("titulados_tsu_inge")
    try:
        c, u, s, warns, errs = _procesar_excel_en_bd(f, semestre_destino=10)
    except ValueError as e:
        messages.error(request, str(e))
        return redirect("titulados_tsu_inge")

    if c: messages.success(request, f"[ING] Se CREARON {c} registro(s).")
    if u: messages.success(request, f"[ING] Se ACTUALIZARON {u} registro(s).")
    if s: messages.info(request, f"[ING] Se saltaron {s} fila(s) vacías o sin programa.")
    if warns: messages.warning(request, " · ".join(warns[:6]) + (" ..." if len(warns) > 6 else ""))
    if errs: messages.error(request, " · ".join(errs[:6]) + (" ..." if len(errs) > 6 else ""))
    return redirect("titulados_tsu_inge")

# =================== VISTA PRINCIPAL ===================
def titulados_tsu_inge_view(request):
    """
    Página: 'Titulados por TSU e Ingeniería'
    Filtros GET:
      - nivel=TSU|ING (default TSU)
      - anio=...     (múltiple)
      - programa=... (múltiple, por NOMBRE exacto)
      - buscar=...   (opcional, por NOMBRE icontains)
    """
    nivel, qs, qs_named = _filtrar_queryset_base(request)

    # Paginación
    try:
        per_page = max(1, min(500, int(request.GET.get("per_page", 50))))
    except Exception:
        per_page = 50
    page_number = request.GET.get("page", 1)
    paginator = Paginator(qs_named, per_page)
    registros_page = paginator.get_page(page_number)

    # Catálogos (tras filtros)
    cat_anios = qs.values_list("anio_ingreso", flat=True).distinct().order_by("anio_ingreso")
    cat_programas = qs_named.values_list("programa_nombre", flat=True).distinct().order_by("programa_nombre")

    # (opcionales) catálogos globales
    programas_antiguos = ProgramaEducativoAntiguo.objects.only("id", "nombre").order_by("nombre")
    programas_nuevos   = ProgramaEducativoNuevo.objects.only("id", "nombre").order_by("nombre")

    # Datos para gráficas
    datos_qs = (
        qs_named.annotate(
            total_tit=Coalesce(F("titulados_hombres"), V(0)) + Coalesce(F("titulados_mujeres"), V(0)),
            total_dgp_calc=Coalesce(F("registrados_dgp_h"), V(0)) + Coalesce(F("registrados_dgp_m"), V(0)),
        )
        .values(
            "programa_nombre", "anio_ingreso", "anio_egreso",
            "titulados_hombres", "titulados_mujeres",
            "registrados_dgp_h", "registrados_dgp_m",
            "total_tit", "total_dgp_calc",
        )
    )

    datos_json = [
        {
            "programa": r["programa_nombre"] or "SIN PROGRAMA",
            "anio_ingreso": int(r["anio_ingreso"] or 0),
            "anio_egreso": int(r["anio_egreso"] or 0),
            "titulados_h": int(r["titulados_hombres"] or 0),
            "titulados_m": int(r["titulados_mujeres"] or 0),
            "dgp_h": int(r["registrados_dgp_h"] or 0),
            "dgp_m": int(r["registrados_dgp_m"] or 0),
            "total_titulados": int(r["total_tit"] or 0),
            "total_dgp": int(r["total_dgp_calc"] or 0),
        }
        for r in datos_qs
    ]

    return render(request, "titulados_tsu_inge.html", {
        "nivel": nivel,
        "registros": registros_page,
        "paginator": paginator,
        "page_obj": registros_page,
        "anios": cat_anios,
        "programas_disponibles": list(cat_programas),
        "programas_antiguos": programas_antiguos,
        "programas_nuevos": programas_nuevos,
        "datos_json": datos_json,
    })
