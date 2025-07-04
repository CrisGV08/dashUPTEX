from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from api.models import GeneracionCarrera, ProgramaEducativoAntiguo, ProgramaEducativoNuevo
import openpyxl
import json
import io

def titulados_historicos_view(request):
    registros = GeneracionCarrera.objects.all().order_by('anio_ingreso')

    if request.method == 'POST' and 'guardar' in request.POST:
        try:
            id_registro = int(request.POST.get('guardar'))
            registro = GeneracionCarrera.objects.get(id=id_registro)
            registro.anio_ingreso = int(request.POST.get(f'anio_ingreso_{id_registro}', 0))
            registro.anio_egreso = int(request.POST.get(f'anio_egreso_{id_registro}', 0))
            registro.genero = request.POST.get(f'genero_{id_registro}', '')
            registro.titulados = int(request.POST.get(f'titulados_{id_registro}', 0))
            registro.registrados_dgp = int(request.POST.get(f'registrados_{id_registro}', 0))

            # Calcula tasa
            if registro.titulados > 0:
                registro.tasa_titulacion = round((registro.registrados_dgp / registro.titulados) * 100, 2)
            else:
                registro.tasa_titulacion = 0.0

            registro.save()
            messages.success(request, "✅ Registro actualizado correctamente.")
            return redirect('titulados_historicos')
        except Exception as e:
            messages.error(request, f"❌ Error al actualizar: {e}")

    # Datos auxiliares
    anios = GeneracionCarrera.objects.values_list('anio_ingreso', flat=True).distinct().order_by('anio_ingreso')
    programas_antiguos = ProgramaEducativoAntiguo.objects.all()
    programas_nuevos = ProgramaEducativoNuevo.objects.all()

    # JSON para gráficas
    datos_json = []
    for r in registros:
        nombre_programa = r.programa_antiguo.nombre if r.programa_antiguo else (r.programa_nuevo.nombre if r.programa_nuevo else "Sin Programa")
        datos_json.append({
            'programa': nombre_programa,
            'anio_ingreso': r.anio_ingreso,
            'anio_egreso': r.anio_egreso,
            'genero': r.genero,
            'titulados': r.titulados,
            'registrados_dgp': r.registrados_dgp,
            'tasa_titulacion': r.tasa_titulacion,
        })

    context = {
        'registros': registros,
        'anios': anios,
        'programas_antiguos': programas_antiguos,
        'programas_nuevos': programas_nuevos,
        'datos_json': json.dumps(datos_json)
    }

    return render(request, 'titulados_historicos.html', context)


def descargar_plantilla_titulados(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Titulados Históricos"
    ws.append(["Programa Antiguo", "Programa Nuevo", "Año Ingreso", "Año Egreso", "Género", "Titulados", "Registrados DGP"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=plantilla_titulados_historicos.xlsx'
    return response


def subir_excel_titulados(request):
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo = request.FILES['archivo_excel']
        try:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            filas = list(ws.iter_rows(min_row=2, values_only=True))

            nuevos, omitidos = 0, 0
            for fila in filas:
                if len(fila) < 7:
                    omitidos += 1
                    continue

                prog_antiguo, prog_nuevo, anio_in, anio_out, genero, titulados, registrados = fila

                if not anio_in or titulados is None or registrados is None:
                    omitidos += 1
                    continue

                pa = ProgramaEducativoAntiguo.objects.filter(nombre=prog_antiguo).first() if prog_antiguo else None
                pn = ProgramaEducativoNuevo.objects.filter(nombre=prog_nuevo).first() if prog_nuevo else None

                if not (pa or pn):
                    omitidos += 1
                    continue

                registro, created = GeneracionCarrera.objects.get_or_create(
                    programa_antiguo=pa,
                    programa_nuevo=pn,
                    anio_ingreso=anio_in,
                    anio_egreso=anio_out,
                    genero=genero,
                    defaults={
                        'titulados': titulados or 0,
                        'registrados_dgp': registrados or 0
                    }
                )

                if not created:
                    registro.titulados = titulados or 0
                    registro.registrados_dgp = registrados or 0

                if registro.titulados > 0:
                    registro.tasa_titulacion = round((registro.registrados_dgp / registro.titulados) * 100, 2)
                else:
                    registro.tasa_titulacion = 0.0

                registro.save()
                nuevos += 1

            msg = f"✅ Se cargaron {nuevos} registros."
            if omitidos > 0:
                msg += f" ⚠️ {omitidos} filas omitidas."
            messages.success(request, msg)

        except Exception as e:
            messages.error(request, f"❌ Error al procesar el archivo: {e}")

    return redirect('titulados_historicos')
